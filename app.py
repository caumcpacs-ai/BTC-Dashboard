
import os
import sqlite3
import datetime
import openpyxl
from flask import Flask, render_template, request, jsonify

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = 'ultrasound_2026'

DATA_DIR  = os.path.join(BASE_DIR, 'data')
DB_PATH   = os.path.join(DATA_DIR, 'joined.db')
REF_PATH  = os.path.join(DATA_DIR, 'ref_file.xlsx')
DATA_PATH = os.path.join(DATA_DIR, 'data_file.xlsx')

os.makedirs(DATA_DIR, exist_ok=True)


# ── SQLite 유틸 ───────────────────────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ── Excel 파싱 ────────────────────────────────────────────────────────────────

def load_ref(path):
    """기준파일 → {코드: 상태} 딕셔너리"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    ref = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # No(0), 상태(1), 코드(2)
        status = row[1]
        code   = row[2]
        if code and status:
            ref[str(code).strip()] = str(status).strip()
    wb.close()
    return ref


def load_data_columns(path):
    """데이터파일 헤더(1행) 반환"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    return list(headers)


def safe_str(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.time):
        return v.strftime('%H:%M:%S')
    return str(v)


def join_and_build_db():
    """기준파일 + 데이터파일 조인 → joined.db 생성"""
    ref_map = load_ref(REF_PATH)

    wb = openpyxl.load_workbook(DATA_PATH, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        raise ValueError('데이터파일이 비어있습니다.')

    headers = list(rows[0])
    data_rows = rows[1:]

    try:
        code_idx = headers.index('처방코드')
    except ValueError:
        raise ValueError("데이터파일에 '처방코드' 컬럼이 없습니다.")

    final_cols = ['상태'] + headers

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = get_conn()
    cur  = conn.cursor()

    col_defs = ', '.join(f'"{c}" TEXT' for c in final_cols)
    cur.execute(f'CREATE TABLE joined_data ({col_defs})')

    placeholders = ', '.join(['?'] * len(final_cols))
    sql = f'INSERT INTO joined_data VALUES ({placeholders})'

    for row in data_rows:
        code   = str(row[code_idx]).strip() if row[code_idx] else ''
        status = ref_map.get(code, '')
        values = [status] + [safe_str(v) for v in row]
        cur.execute(sql, values)

    conn.commit()
    conn.close()

    return len(data_rows)


# ── 파일 상태 헬퍼 ────────────────────────────────────────────────────────────

def file_info(path):
    if os.path.exists(path):
        size = os.path.getsize(path)
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path))
        return {
            'exists': True,
            'name': os.path.basename(path),
            'size': f'{size/1024:.1f} KB',
            'mtime': mtime.strftime('%Y-%m-%d %H:%M'),
        }
    return {'exists': False}


# ── 라우트 ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    ref_info  = file_info(REF_PATH)
    data_info = file_info(DATA_PATH)
    db_info   = file_info(DB_PATH)

    stats = None
    if db_info['exists']:
        stats = get_stats()

    return render_template('index.html',
                           ref_info=ref_info,
                           data_info=data_info,
                           db_info=db_info,
                           stats=stats)


@app.route('/upload/<filetype>', methods=['POST'])
def upload(filetype):
    if filetype not in ('ref', 'data'):
        return jsonify({'success': False, 'message': '잘못된 요청입니다.'})

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'xlsx 파일만 업로드 가능합니다.'})

    save_path = REF_PATH if filetype == 'ref' else DATA_PATH
    file.save(save_path)

    label = '기준파일' if filetype == 'ref' else '데이터파일'
    info  = file_info(save_path)
    return jsonify({'success': True, 'message': f'{label} 저장 완료', 'info': info})


@app.route('/delete/<filetype>', methods=['POST'])
def delete(filetype):
    if filetype not in ('ref', 'data'):
        return jsonify({'success': False, 'message': '잘못된 요청입니다.'})

    path  = REF_PATH if filetype == 'ref' else DATA_PATH
    label = '기준파일' if filetype == 'ref' else '데이터파일'

    if os.path.exists(path):
        os.remove(path)
        if os.path.exists(DB_PATH):
            os.remove(DB_PATH)
        return jsonify({'success': True, 'message': f'{label} 삭제 완료'})
    return jsonify({'success': False, 'message': f'{label}이 없습니다.'})


@app.route('/join', methods=['POST'])
def join():
    if not os.path.exists(REF_PATH):
        return jsonify({'success': False, 'message': '기준파일이 없습니다.'})
    if not os.path.exists(DATA_PATH):
        return jsonify({'success': False, 'message': '데이터파일이 없습니다.'})
    try:
        count = join_and_build_db()
        return jsonify({'success': True, 'message': f'조인 완료 — {count:,}건 저장', 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'message': f'오류: {e}'})


@app.route('/api/stats')
def api_stats():
    if not os.path.exists(DB_PATH):
        return jsonify({'error': 'DB 없음'})
    return jsonify(get_stats())


def get_stats():
    conn = get_conn()
    cur  = conn.cursor()

    def q(sql):
        cur.execute(sql)
        return cur.fetchall()

    status_rows = q('SELECT "상태", COUNT(*) FROM joined_data GROUP BY "상태" ORDER BY COUNT(*) DESC')

    try:
        monthly_rows = q('''
            SELECT "Year", "Month", "상태", COUNT(*)
            FROM joined_data
            WHERE "Year" IS NOT NULL AND "Month" IS NOT NULL
            GROUP BY "Year", "Month", "상태"
            ORDER BY "Year", "Month", "상태"
        ''')
    except Exception:
        monthly_rows = []

    try:
        section_rows = q('SELECT "Section", "상태", COUNT(*) FROM joined_data WHERE "Section" IS NOT NULL GROUP BY "Section", "상태" ORDER BY COUNT(*) DESC')
    except Exception:
        section_rows = []

    try:
        dept_rows = q('SELECT "처방과", COUNT(*) FROM joined_data WHERE "처방과" IS NOT NULL GROUP BY "처방과" ORDER BY COUNT(*) DESC LIMIT 10')
    except Exception:
        dept_rows = []

    try:
        inout_rows = q('SELECT "입원외래", COUNT(*) FROM joined_data WHERE "입원외래" IS NOT NULL GROUP BY "입원외래"')
    except Exception:
        inout_rows = []

    # Section별 월별 (ED→Thyroid, MAM→Breast)
    try:
        sm_rows = q('''
            SELECT "Year", "Month", "Section", "상태", COUNT(*)
            FROM joined_data
            WHERE "Section" IN ('ED', 'MAM')
              AND "Year" IS NOT NULL AND "Month" IS NOT NULL
            GROUP BY "Year", "Month", "Section", "상태"
            ORDER BY "Year", "Month", "Section"
        ''')
    except Exception:
        sm_rows = []

    cur.execute('SELECT COUNT(*) FROM joined_data')
    total = cur.fetchone()[0]

    conn.close()

    return {
        'total': total,
        'status':  [{'label': r[0] or '미분류', 'count': r[1]} for r in status_rows],
        'monthly': [{'year': r[0], 'month': r[1], 'status': r[2] or '미분류', 'count': r[3]} for r in monthly_rows],
        'section': [{'section': r[0], 'status': r[1] or '미분류', 'count': r[2]} for r in section_rows],
        'dept':    [{'dept': r[0], 'count': r[1]} for r in dept_rows],
        'inout':   [{'label': r[0], 'count': r[1]} for r in inout_rows],
        'section_monthly': [{'year': r[0], 'month': r[1], 'section': r[2], 'status': r[3] or '', 'count': r[4]} for r in sm_rows],
    }


if __name__ == '__main__':
    app.run(port=5000)
